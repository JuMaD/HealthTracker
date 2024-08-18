import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from fpdf import FPDF
from io import BytesIO
import os

def log_function_call(func):
    import logging
    from functools import wraps

    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

    @wraps(func)
    def wrapper(*args, **kwargs):
        logging.info(f"Calling function '{func.__name__}' with arguments {args} and keyword arguments {kwargs}")
        result = func(*args, **kwargs)
        logging.info(f"Function '{func.__name__}' returned {result}")
        return result
    return wrapper

@log_function_call
def import_sanitize(filename):
    df = pd.read_csv(filename, parse_dates=['Datum'], date_parser=lambda x: pd.to_datetime(x, format='%m/%d/%y', errors='coerce'))
    grouped = df.groupby('Bezeichnung')
    for name, group in grouped:
        df.loc[group.index, 'unterer Grenzwert'] = group['unterer Grenzwert'].ffill().bfill()
        df.loc[group.index, 'oberer Grenzwert'] = group['oberer Grenzwert'].ffill().bfill()
    return df, df.groupby('Bezeichnung')

@log_function_call
def save_plots(df, grouped, plots_dir='plots'):
    if not os.path.exists(plots_dir):
        os.makedirs(plots_dir)
    for name, group in grouped:
        if len(group) > 1:
            sorted_group = group.sort_values('Datum')
            overall_lower_bound = sorted_group['unterer Grenzwert'].min() if sorted_group['unterer Grenzwert'].notna().any() else None
            overall_upper_bound = sorted_group['oberer Grenzwert'].max() if sorted_group['oberer Grenzwert'].notna().any() else None
            plt.figure(figsize=(10, 5))
            plt.plot(sorted_group['Datum'], sorted_group['Wert'], marker='o', label=f'{name}')
            if overall_lower_bound is not None and overall_upper_bound is not None:
                plt.fill_between(sorted_group['Datum'], overall_lower_bound, overall_upper_bound, color='gray', alpha=0.3, label='Normal Range')
            elif overall_lower_bound is not None:
                plt.axhline(y=overall_lower_bound, color='gray', linestyle='--', label='Lower Bound')
            elif overall_upper_bound is not None:
                plt.axhline(y=overall_upper_bound, color='gray', linestyle='--', label='Upper Bound')
            plt.title(f'{name} over Time')
            plt.xlabel('Date')
            plt.ylabel(f'{name} ({sorted_group["Einheit"].iloc[0]})')
            plt.legend()
            plt.grid(True)
            plt.xticks(rotation=45)
            plt.tight_layout()
            filename = os.path.join(plots_dir, f'{sanitize_filename(name)}.png')
            plt.savefig(filename)
            plt.close()

@log_function_call
def export_to_excel(df, grouped, filename='lab_results_report.xlsx', selected_values=None):
    df['Datum'] = pd.to_datetime(df['Datum'], format='%m/%d/%y', errors='coerce')
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "All Data"
    df_sorted = df.sort_values(by=['Bezeichnung', 'Datum'], ascending=[True, False])
    for row in dataframe_to_rows(df_sorted, index=False, header=True):
        ws_data.append(row)
    ws_plots = wb.create_sheet(title="Selected Graphs")
    row_offset = 2
    col_offset = 1
    max_columns = 2
    for i, name in enumerate(selected_values):
        if name in grouped.groups:
            group = grouped.get_group(name)
            if len(group) > 1:
                graph_path = os.path.join('plots', f'{sanitize_filename(name)}.png')
                if os.path.exists(graph_path):
                    img = Image(graph_path)
                    row = row_offset + (i // max_columns) * 20
                    col = col_offset + (i % max_columns) * 8
                    cell_position = ws_plots.cell(row=row, column=col).coordinate
                    ws_plots.add_image(img, cell_position)
    ws_current = wb.create_sheet(title="Current Data")
    latest_entries = df.groupby('Bezeichnung').apply(lambda x: x.nlargest(1, 'Datum')).reset_index(drop=True)
    for row in dataframe_to_rows(latest_entries, index=False, header=True):
        ws_current.append(row)
    wb.save(filename)

@log_function_call
def generate_pdf_report(df, selected_values, filename='lab_results_report.pdf', plots_dir='plots'):
    df['Datum'] = pd.to_datetime(df['Datum'], format='%m/%d/%y', errors='coerce')
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, 'Most Current Data', 0, 1, 'C')
    latest_entries = df.groupby('Bezeichnung').apply(lambda x: x.nlargest(1, 'Datum')).reset_index(drop=True)
    latest_entries = latest_entries.drop(columns=['unterer Grenzwert', 'oberer Grenzwert']).fillna('')
    latest_entries['Datum'] = latest_entries['Datum'].dt.strftime('%m/%d/%y')
    pdf.set_font('Arial', 'B', 10)
    for col in latest_entries.columns:
        pdf.cell(40, 10, col, 1)
    pdf.ln()
    pdf.set_font('Arial', '', 10)
    for index, row in latest_entries.iterrows():
        for item in row:
            pdf.cell(40, 10, str(item), 1)
        pdf.ln()
    graph_counter = 0
    for name in selected_values:
        graph_path = os.path.join(plots_dir, f'{sanitize_filename(name)}.png')
        if os.path.exists(graph_path):
            if graph_counter % 2 == 0:
                pdf.add_page()
                y_offset = 30
            else:
                y_offset = 150
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, f'{name} Graph', 0, 1, 'C')
            pdf.image(graph_path, x=10, y=y_offset, w=190)
            graph_counter += 1
    pdf.output(filename)

@log_function_call
def sanitize_filename(name):
    return name.replace(" ", "_").replace("/", "-").lower()
